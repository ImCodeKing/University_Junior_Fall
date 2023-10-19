import os
import openpyxl

# 获取文件夹中的所有文件名
folder_path = 'D:/李昭阳/大学/University_Junior_Fall/机器人/HW2/image_of_boa/image_of_boa'  # 将此路径替换为你的文件夹路径
file_names = os.listdir(folder_path)

# 创建一个新的Excel工作簿
workbook = openpyxl.Workbook()
worksheet = workbook.active

# 遍历文件名列表并将结果写入Excel表格
for index, file_name in enumerate(file_names, start=1):
    # 去除文件拓展名并用 "-" 分割
    name_without_extension = os.path.splitext(file_name)[0]
    name_parts = name_without_extension.split('-')

    # 将结果写入Excel表格的不同列
    worksheet.cell(row=index, column=1, value=file_name)  # 原始文件名
    for col, part in enumerate(name_parts, start=2):
        worksheet.cell(row=index, column=col, value=part)

# 保存Excel文件
excel_file_path = 'D:/李昭阳/大学/University_Junior_Fall/机器人/HW2/image_of_boa/output.xlsx'  # 将此路径替换为你想要保存的Excel文件路径
workbook.save(excel_file_path)

# 关闭工作簿
workbook.close()